"""Extract plain text from PDF, DOCX, and XLSX files."""

from __future__ import annotations

from pathlib import Path

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx"}


class UnsupportedFileTypeError(ValueError):
    """Raised when a file extension is not supported."""


def parse_file(path: Path) -> str:
    """Dispatch to the correct parser based on file extension."""
    ext = path.suffix.lower()
    if ext == ".pdf":
        return parse_pdf(path)
    if ext == ".docx":
        return parse_docx(path)
    if ext == ".xlsx":
        return parse_xlsx(path)
    raise UnsupportedFileTypeError(
        f"Unsupported file type '{ext}'. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
    )


def parse_pdf(path: Path) -> str:
    """Extract plain text from a PDF using PyMuPDF."""
    try:
        import pymupdf  # type: ignore[import-untyped]
    except ImportError:
        raise ImportError("pymupdf is required: uv pip install 'ai-assistant[docs]'")

    doc = pymupdf.open(str(path))
    pages: list[str] = []
    for page in doc:
        text = page.get_text()
        if text.strip():
            pages.append(text)
    doc.close()

    result = "\n\n".join(pages)
    if not result.strip():
        raise ValueError(
            f"{path.name}: no text extracted. The PDF may be image-only (scanned). "
            "OCR is not supported in v1."
        )
    return result


def parse_docx(path: Path) -> str:
    """Extract plain text from a Word document."""
    try:
        import docx  # type: ignore[import-untyped]
    except ImportError:
        raise ImportError("python-docx is required: uv pip install 'ai-assistant[docs]'")

    doc = docx.Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def parse_xlsx(path: Path) -> str:
    """Extract text from an Excel workbook, sheet by sheet."""
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        raise ImportError("openpyxl is required: uv pip install 'ai-assistant[docs]'")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            line = "\t".join(cells).rstrip()
            if line.strip():
                rows.append(line)
        if rows:
            sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    wb.close()
    return "\n\n".join(sheets)


def iter_supported_files(directory: Path) -> list[Path]:
    """Recursively find all supported files in a directory."""
    files: list[Path] = []
    for path in sorted(directory.rglob("*")):
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(path)
    return files
